import io
import os
import sys
import unittest
from pathlib import Path
from urllib.parse import unquote_plus
from unittest.mock import patch

from docx import Document
from openpyxl import load_workbook


APP_DIR = Path(__file__).resolve().parents[1] / "compras_app"
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))
os.environ["SUPRIMENTOS_FILE_LOG"] = "0"

import app as app_module  # noqa: E402
import gerar_oc  # noqa: E402
import supabase_data  # noqa: E402


class PurchaseTransitTests(unittest.TestCase):
    def setUp(self):
        app_module.app.config.update(TESTING=True)
        self.client = app_module.app.test_client()

    def test_live_transit_uses_shared_order_lines_and_pending_balance(self):
        orders = [{
            "id": "order-1",
            "numero_oc": "2801",
            "categoria": "GERAL",
            "fornecedor_nome": "Fornecedor A",
            "status": "EMITIDA",
            "data_emissao": "2026-08-03",
            "data_necessidade": "2999-08-01",
            "destino": "ESTOQUE",
            "technical_status": None,
        }]
        lines = [
            {
                "id": "line-1",
                "purchase_order_id": "order-1",
                "numero_linha": 1,
                "sku_codigo": "SKU-1",
                "descricao_original": "Item em trânsito",
                "unidade": "pc",
                "quantidade_pedida": "10",
                "quantidade_recebida": "4",
                "destino": "O.S. 3100",
                "data_necessidade": "2999-08-15",
                "status": "PARCIALMENTE_RECEBIDA",
            },
            {
                "id": "line-2",
                "purchase_order_id": "order-1",
                "numero_linha": 2,
                "sku_codigo": "SKU-2",
                "descricao_original": "Item já completo",
                "unidade": "pc",
                "quantidade_pedida": 2,
                "quantidade_recebida": 2,
                "data_necessidade": "2999-08-20",
                "status": "PARCIALMENTE_RECEBIDA",
            },
        ]
        calls = []

        def fake_all_rows(table, **kwargs):
            calls.append((table, kwargs))
            return orders if table == supabase_data.PURCHASE_ORDERS_TABLE else lines

        with patch.object(supabase_data, "_all_rows", side_effect=fake_all_rows):
            rows = supabase_data.carregar_compras_transito(force=True)

        self.assertEqual(1, len(rows))
        self.assertEqual(6, rows[0]["quantidade_pendente"])
        self.assertEqual("2999-08-15", rows[0]["data_necessidade"])
        self.assertEqual("A VENCER", rows[0]["situacao_transito"])
        self.assertEqual("O.S. 3100", rows[0]["destino"])
        order_query = calls[0][1]["extra_query"]
        self.assertIn(
            ("or", "(technical_status.is.null,technical_status.neq.CONCLUIDA)"),
            order_query,
        )

    def test_purchase_form_keeps_a_separate_remittance_date_per_line(self):
        with (
            patch.object(app_module, "login_enabled", return_value=False),
            patch.object(app_module, "atualizar_skus_automatico", return_value={}),
            patch.object(app_module, "carregar_fornecedores", return_value={}),
            patch.object(app_module, "carregar_produtos", return_value={}),
            patch.object(app_module, "carregar_os_componentes", return_value={}),
            patch.object(app_module, "proximo_numero_oc", return_value=2801),
            patch.object(app_module, "registrar_historico") as register,
            patch.object(app_module, "gerar_word") as generate,
        ):
            response = self.client.post("/gerar_oc", data={
                "acao": "salvar",
                "oc_submit_token": "dates-by-line",
                "fornecedor": "Fornecedor",
                "previsao": "2026-08-10",
                "codigo[]": ["SKU-1", "SKU-2"],
                "descricao[]": ["Item 1", "Item 2"],
                "unidade[]": ["UN", "UN"],
                "qtd[]": ["2", "3"],
                "data_necessidade[]": ["2026-08-10", "2026-09-15"],
                "valor[]": ["10", "20"],
                "desconto[]": ["0", "0"],
                "frete": "0",
            })

        self.assertEqual(302, response.status_code)
        generate.assert_not_called()
        saved_items = register.call_args.kwargs["itens"]
        self.assertEqual(
            ["2026-08-10", "2026-09-15"],
            [item["data_necessidade"] for item in saved_items],
        )
        # A data oficial do cabeçalho é a última remessa prevista do pedido.
        self.assertEqual("2026-09-15", register.call_args.args[2]["previsao"])

    def test_new_purchase_with_retroactive_remittance_is_not_emitted(self):
        yesterday = (app_module._operational_today() - app_module.timedelta(days=1)).isoformat()
        with (
            patch.object(app_module, "login_enabled", return_value=False),
            patch.object(app_module, "atualizar_skus_automatico", return_value={}),
            patch.object(app_module, "carregar_fornecedores", return_value={}),
            patch.object(app_module, "carregar_produtos", return_value={}),
            patch.object(app_module, "registrar_historico") as register,
            patch.object(app_module, "gerar_word") as generate,
        ):
            response = self.client.post("/gerar_oc", data={
                "fornecedor": "Fornecedor",
                "codigo[]": ["SKU-1"],
                "descricao[]": ["Item 1"],
                "unidade[]": ["UN"],
                "qtd[]": ["2"],
                "data_necessidade[]": [yesterday],
                "valor[]": ["10"],
                "desconto[]": ["0"],
                "frete": "0",
            })

        self.assertEqual(302, response.status_code)
        self.assertIn("datas de remessa retroativas", unquote_plus(response.headers["Location"]))
        register.assert_not_called()
        generate.assert_not_called()

    def test_purchase_draft_may_keep_retroactive_date_until_emission(self):
        yesterday = (app_module._operational_today() - app_module.timedelta(days=1)).isoformat()
        with (
            patch.object(app_module, "login_enabled", return_value=False),
            patch.object(app_module, "atualizar_skus_automatico", return_value={}),
            patch.object(app_module, "carregar_fornecedores", return_value={}),
            patch.object(app_module, "carregar_produtos", return_value={}),
            patch.object(app_module, "carregar_os_componentes", return_value={}),
            patch.object(app_module, "proximo_numero_oc", return_value=2801),
            patch.object(app_module, "registrar_historico") as register,
            patch.object(app_module, "gerar_word") as generate,
        ):
            response = self.client.post("/gerar_oc", data={
                "acao": "salvar",
                "oc_submit_token": "retroactive-draft",
                "fornecedor": "Fornecedor",
                "codigo[]": ["SKU-1"],
                "descricao[]": ["Item 1"],
                "unidade[]": ["UN"],
                "qtd[]": ["2"],
                "data_necessidade[]": [yesterday],
                "valor[]": ["10"],
                "desconto[]": ["0"],
                "frete": "0",
            })

        self.assertEqual(302, response.status_code)
        register.assert_called_once()
        generate.assert_not_called()

    def test_erp_sync_publishes_each_line_remittance_date(self):
        items = [
            {"codigo": "SKU-1", "descricao": "Item 1", "unidade": "UN", "qtd": 1, "valor": 10, "data_necessidade": "2026-08-10"},
            {"codigo": "SKU-2", "descricao": "Item 2", "unidade": "UN", "qtd": 1, "valor": 20, "data_necessidade": "2026-09-15"},
        ]
        with (
            patch.object(app_module, "erp_feature_enabled", return_value=True),
            patch.object(app_module, "_erp_stock_request", return_value={"id": "order-1"}) as request_erp,
        ):
            app_module._sync_emitted_legacy_oc_to_erp(
                {"id": "doc-1", "data_criacao": "2026-08-03"},
                {"previsao": "2026-08-10", "destino": "ESTOQUE"},
                items,
                "2801",
                "Fornecedor",
            )

        payload = request_erp.call_args.args[2]
        self.assertEqual(
            ["2026-08-10", "2026-09-15"],
            [line["data_necessidade"] for line in payload["lines"]],
        )

    def test_emitted_purchase_document_shows_remittance_date_on_each_line(self):
        path = gerar_oc.gerar_word(
            "2801",
            "Fornecedor",
            {
                "previsao": "2026-08-10",
                "total_itens": 10,
                "total_pedido": 10,
            },
            [{
                "codigo": "SKU-1",
                "descricao": "Item 1",
                "unidade": "UN",
                "qtd": 5.0,
                "valor": 10,
                "desconto": 0,
                "total": 10,
                "data_necessidade": "2026-09-15",
            }],
            incluir_composicao=False,
            componentes={},
        )
        try:
            document = Document(path)
            product_table = next(
                table
                for table in document.tables
                if any(
                    "DATA DE REMESSA" in " ".join(cell.text.split()).upper()
                    for cell in table.rows[0].cells
                )
            )
            headers = [" ".join(cell.text.split()).upper() for cell in product_table.rows[0].cells]
            remittance_column = headers.index("DATA DE REMESSA")
            body = product_table.rows[1].cells
            self.assertEqual("Item 1", " ".join(body[1].text.split()))
            self.assertEqual("5", body[3].text.strip())
            self.assertEqual("15/09/2026", body[remittance_column].text.strip())
        finally:
            generated = Path(path)
            generated.unlink(missing_ok=True)
            generated.parent.rmdir()

    def test_document_quantity_format_removes_only_trailing_zeroes(self):
        self.assertEqual("5", gerar_oc._format_quantity(5.0))
        self.assertEqual("2,5", gerar_oc._format_quantity("2,50"))

    def test_transit_api_and_export_share_the_same_projection(self):
        rows = [{
            "purchase_order_id": "order-1",
            "purchase_order_line_id": "line-1",
            "numero_oc": "2801",
            "numero_linha": 1,
            "fornecedor_nome": "Fornecedor",
            "sku_codigo": "SKU-1",
            "descricao_original": "Item",
            "unidade": "pc",
            "quantidade_pedida": 10,
            "quantidade_recebida": 4,
            "quantidade_pendente": 6,
            "data_necessidade": "2026-08-10",
            "situacao_transito": "A VENCER",
            "dias_para_remessa": 7,
            "destino": "ESTOQUE",
            "status": "PARCIALMENTE_RECEBIDA",
        }]
        with (
            patch.object(app_module, "login_enabled", return_value=False),
            patch.object(app_module, "erp_feature_enabled", return_value=True),
            patch.object(app_module, "can", return_value=True),
            patch.object(app_module.supabase_data, "carregar_compras_transito", return_value=rows),
        ):
            api_response = self.client.get("/api/erp/purchase-orders/transit")
            export_response = self.client.get("/erp/relatorios/compras-transito.xlsx")

        self.assertEqual(200, api_response.status_code)
        self.assertEqual(6, api_response.get_json()["metrics"]["quantidade_pendente"])
        self.assertEqual(200, export_response.status_code)
        workbook = load_workbook(io.BytesIO(export_response.data), read_only=True)
        sheet = workbook["Transito pendente"]
        values = list(sheet.iter_rows(values_only=True))
        self.assertIn("Data necessidade / remessa", values[0])
        self.assertIn("Situacao do transito", values[0])
        self.assertEqual("2801", str(values[1][0]))
        self.assertEqual(6, values[1][8])
        workbook.close()
        export_response.close()

    def test_management_screen_exposes_transit_panel_and_export(self):
        template = (APP_DIR / "templates" / "erp_ordens_compra.html").read_text(encoding="utf-8")
        self.assertIn("Trânsito pendente por remessa", template)
        self.assertIn("/erp/relatorios/compras-transito.xlsx", template)
        self.assertIn("/api/erp/purchase-orders/transit", template)
        with (
            patch.object(app_module, "login_enabled", return_value=False),
            patch.object(app_module, "erp_feature_enabled", return_value=True),
            patch.object(app_module, "can", return_value=True),
        ):
            response = self.client.get("/erp/ordens-compra")
        self.assertEqual(200, response.status_code)
        self.assertIn("Trânsito pendente por remessa".encode(), response.data)


if __name__ == "__main__":
    unittest.main()
