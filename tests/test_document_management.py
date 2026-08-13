import io
import json
import os
import sys
import tempfile
import unittest
import copy
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook


APP_DIR = Path(__file__).resolve().parents[1] / "compras_app"
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))
os.environ["SUPRIMENTOS_FILE_LOG"] = "0"

import app as app_module
import supabase_data


class DocumentManagementTests(unittest.TestCase):
    def test_reissue_preserves_direct_billing_supplier_from_history(self):
        previous_document = {
            "itens": [
                {"codigo": "10100001", "fornecedor": "INSTALL TECH"},
            ],
            "composicao": [
                {
                    "item": "10100001",
                    "codigo": "10100001",
                    "descricao": "FATURAMENTO DIRETO - INSTALACAO",
                    "fornecedor": "INSTALL TECH",
                    "tipo_requisicao": "FATURAMENTO DIRETO",
                    "setor": "F.D",
                },
            ],
        }

        suppliers = app_module._fornecedores_faturamento_direto_historicos(previous_document)
        composition = app_module._preservar_fornecedor_faturamento_direto(
            [
                {
                    "item": "10100001",
                    "codigo": "10100001",
                    "descricao": "FATURAMENTO DIRETO - INSTALACAO",
                    "tipo_requisicao": "FATURAMENTO DIRETO",
                    "setor": "F.D",
                    "fornecedor": "",
                }
            ],
            [{"codigo": "10100001", "fornecedor": ""}],
            suppliers,
        )

        self.assertEqual("INSTALL TECH", composition[0]["fornecedor"])

    def test_manual_composition_parser_preserves_supplier(self):
        with app_module.app.test_request_context(
            "/gerar_os",
            method="POST",
            data={
                "os_comp_item[]": "10100001",
                "os_comp_codigo[]": "10100001",
                "os_comp_descricao[]": "FATURAMENTO DIRETO - INSTALACAO",
                "os_comp_unidade[]": "un",
                "os_comp_qtd[]": "1",
                "os_comp_level[]": "0",
                "os_comp_setor[]": "F.D",
                "os_comp_setor_manual[]": "0",
                "os_comp_fornecedor[]": "INSTALL TECH",
            },
        ):
            composition = app_module._parse_os_composition_form(app_module.request.form)

        self.assertEqual("INSTALL TECH", composition[0]["fornecedor"])

    def test_document_normalization_preserves_layout_reference(self):
        document = supabase_data.normalizar_documento({
            "tipo": "os",
            "numero": "3119",
            "layout_arquivo_id": "d468c8aa-7cab-4b8d-84cf-0d5aef6c0f82",
        })

        self.assertEqual("d468c8aa-7cab-4b8d-84cf-0d5aef6c0f82", document["layout_arquivo_id"])

    def test_os_layout_preview_streams_the_saved_pdf_inline(self):
        layout_id = "d468c8aa-7cab-4b8d-84cf-0d5aef6c0f82"
        with (
            patch.object(app_module, "login_enabled", return_value=False),
            patch.object(app_module.supabase_data, "enabled", return_value=True),
            patch.object(
                app_module.supabase_data,
                "baixar_layout_pdf",
                return_value=(
                    b"%PDF-1.4\nlayout",
                    {"nome_exibicao": "layout-teste.pdf", "mime_type": "application/pdf"},
                ),
            ),
        ):
            response = app_module.app.test_client().get(f"/api/layouts/{layout_id}/arquivo")

        self.assertEqual(200, response.status_code)
        self.assertEqual("application/pdf", response.mimetype)
        self.assertIn("inline", response.headers.get("Content-Disposition", ""))
        self.assertEqual(b"%PDF-1.4\nlayout", response.data)

    def test_layout_catalog_is_available_for_os_selection(self):
        layouts = [
            {
                "id": "d468c8aa-7cab-4b8d-84cf-0d5aef6c0f82",
                "nome_exibicao": "Sprinter padrão.pdf",
                "tamanho_bytes": 1024,
            }
        ]
        with (
            patch.object(app_module, "login_enabled", return_value=False),
            patch.object(app_module.supabase_data, "enabled", return_value=True),
            patch.object(app_module.supabase_data, "listar_layouts", return_value=layouts),
        ):
            response = app_module.app.test_client().get("/api/layouts")

        self.assertEqual(200, response.status_code)
        self.assertTrue(response.json["ok"])
        self.assertEqual(layouts, response.json["layouts"])

    def test_layout_catalog_query_uses_metadata_only(self):
        with patch.object(supabase_data, "_request", return_value=[]) as request_mock:
            self.assertEqual([], supabase_data.listar_layouts())

        method, table = request_mock.call_args.args[:2]
        query = request_mock.call_args.kwargs["query"]
        self.assertEqual("GET", method)
        self.assertEqual("layout_arquivos", table)
        self.assertIn(("select", "id,nome_original,nome_exibicao,mime_type,tamanho_bytes,criado_por,created_at,updated_at"), query)
        self.assertNotIn(("select", "storage_path"), query)

    def test_quantity_parser_preserves_decimal_points_from_os_documents(self):
        self.assertEqual("1", app_module._normalizar_qtd("1.0"))
        self.assertEqual("2", app_module._normalizar_qtd("2.0"))
        self.assertEqual("0.025", app_module._normalizar_qtd("0.025"))
        self.assertEqual("1.064", app_module._normalizar_qtd("1.064"))
        self.assertEqual("1.5", app_module._normalizar_qtd("1,5"))
        self.assertEqual("1234.5", app_module._normalizar_qtd("1.234,5"))
        self.assertEqual(
            app_module._assinatura_os_recalculada({"numero": "1", "itens": [{"qtd": 1}]}),
            app_module._assinatura_os_recalculada({"numero": "1", "itens": [{"qtd": 1.0}]}),
        )

    def test_historical_luminaire_fallback_requires_one_consistent_relation(self):
        documents = [
            {
                "tipo": "os",
                "composicao": [
                    {"item": "40340010", "codigo": "10260092", "qtd": 2},
                    {"item": "40340020", "codigo": "10260092", "qtd": 2},
                ],
            },
            {
                "tipo": "os",
                "composicao": [
                    {"item": "40340010", "codigo": "10260092", "qtd": 2.0},
                    {"item": "40340020", "codigo": "10260095", "qtd": 2},
                    {
                        "item": "40340030",
                        "codigo": "10260095",
                        "qtd": 3,
                        "line_status": "cancelado",
                    },
                ],
            },
        ]

        result = app_module._luminarias_historicas_por_item(documents)

        self.assertEqual({"codigo": "10260092", "qtd": 2.0}, result["40340010"])
        self.assertNotIn("40340020", result)
        self.assertNotIn("40340030", result)

    def test_reconciled_os_uses_historical_luminaire_when_source_has_none(self):
        source = {
            "nome": "OS-3102.docx",
            "data_arquivo": datetime(2026, 7, 24, 16, 26),
            "dados": {
                "os_numero": "3102",
                "cliente": "BELISA",
                "chassis": "VE277925",
                "itens": [{"codigo": "40340010", "qtd": 1}],
                "composicao": [],
            },
        }
        context = {
            "os_produtos": {
                "40340010": {"descricao": "ITEM RAIZ", "unidade": "un"},
                "10260092": {"descricao": "ELETRICA LUMINARIA ALURE 30", "unidade": "un"},
            },
            "produtos_catalogo": {},
            "componentes": {},
            "processos": {},
            "processo_por_item": {},
            "regras_por_gatilho": {},
            "luminarias_por_item": {
                "40340010": {"codigo": "10260092", "qtd": 2},
            },
        }

        with patch.object(app_module, "current_username", return_value="codex"):
            result = app_module._montar_os_reconciliada(source, None, context)

        luminaires = [
            line for line in result["composicao"]
            if line.get("codigo") == "10260092"
        ]
        self.assertEqual(1, len(luminaires))
        self.assertEqual("40340010", luminaires[0]["item"])
        self.assertEqual(2.0, luminaires[0]["qtd"])

    def test_dashboard_documents_exposes_compact_filter_data(self):
        rows = app_module._dashboard_documentos([{
            "id": 42,
            "tipo": "OS",
            "numero": 3090,
            "data_criacao": "2026-07-21T14:30:00",
            "status": "concluido",
            "updated_at": "2026-07-21T15:00:00",
            "dados": {
                "cliente": "Cliente A",
                "chassis": "ABC123",
                "mmv": "MODELO X",
                "campo_pesado": "nao deve ser enviado",
            },
            "itens": [{"codigo": "SKU-1"}, {"codigo": "SKU-2"}],
            "processos": {"CORTE": [{"atividade": "Cortar"}]},
        }])

        self.assertEqual([{
            "id": "42",
            "tipo": "os",
            "numero": "3090",
            "data": "2026-07-21",
            "status": "concluido",
            "nome": "Cliente A",
            "chassis": "ABC123",
            "mmv": "MODELO X",
            "total": 0.0,
            "itens": 2,
            "ordem": "2026-07-21T15:00:00",
        }], rows)

    def test_production_upgrade_is_additive_and_transactional(self):
        migration_path = Path(__file__).resolve().parents[1] / "supabase_suprimentos_gestao_documentos_additive.sql"
        sql = migration_path.read_text(encoding="utf-8")
        executable_sql = "\n".join(
            line for line in sql.splitlines()
            if not line.lstrip().startswith("--")
        ).lower()

        self.assertIn("begin;", executable_sql)
        self.assertIn("commit;", executable_sql)
        self.assertNotIn("delete from", executable_sql)
        self.assertNotIn("truncate", executable_sql)
        self.assertNotIn("drop table", executable_sql)
        self.assertNotIn("insert into public.suprimentos_documentos ", executable_sql)
        self.assertNotIn("update public.suprimentos_documentos ", executable_sql)

    def test_local_history_is_idempotent_by_submission_token(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            history_path = str(Path(tmpdir) / "historico.json")
            Path(history_path).write_text("[]", encoding="utf-8")
            with (
                patch.object(app_module, "HISTORICO_FILE", history_path),
                patch.object(app_module.supabase_data, "enabled", return_value=False),
                app_module.app.test_request_context("/"),
            ):
                app_module.session["suprimentos_user"] = {"id": 7, "username": "paulo"}
                first = app_module.registrar_historico(
                    "oc", 10, {"fornecedor": "Fornecedor A"}, itens=[{"codigo": "1"}],
                    status="rascunho", submit_token="token-unico",
                )
                second = app_module.registrar_historico(
                    "oc", 10, {"fornecedor": "Fornecedor B"}, itens=[{"codigo": "2"}],
                    status="emitido", submit_token="token-unico",
                )
                rows = json.loads(Path(history_path).read_text(encoding="utf-8"))

            self.assertEqual(1, len(rows))
            self.assertEqual(first["id"], second["id"])
            self.assertEqual("Fornecedor B", rows[0]["dados"]["fornecedor"])
            self.assertEqual("emitido", rows[0]["status"])
            self.assertEqual("paulo", rows[0]["atualizado_por"])

    def test_document_history_assigns_line_ids(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            history_path = str(Path(tmpdir) / "historico.json")
            Path(history_path).write_text("[]", encoding="utf-8")
            with (
                patch.object(app_module, "HISTORICO_FILE", history_path),
                patch.object(app_module.supabase_data, "enabled", return_value=False),
                app_module.app.test_request_context("/"),
            ):
                entry = app_module.registrar_historico(
                    "os",
                    77,
                    {"cliente": "Cliente"},
                    itens=[{"codigo": "SKU-1", "descricao": "Item", "qtd": 1}],
                    processos={"CORTE": [{"atividade": "Cortar", "responsavel": "Ana"}]},
                    composicao=[{"item": "SKU-1", "codigo": "COMP-1", "qtd": 1}],
                    status="emitido",
                    submit_token="token-os",
                )

        self.assertTrue(entry["itens"][0]["line_id"].startswith("os-item-"))
        self.assertTrue(entry["processos"]["CORTE"][0]["line_id"].startswith("os-proc-"))
        self.assertTrue(entry["composicao"][0]["line_id"].startswith("os-comp-"))

    def test_bulk_status_upload_updates_line_or_document_by_scope(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            history_path = Path(tmpdir) / "historico.json"
            history_path.write_text(json.dumps([
                {
                    "id": "local-oc",
                    "tipo": "oc",
                    "numero": "100",
                    "data_criacao": "2026-07-19",
                    "status": "emitido",
                    "dados": {"fornecedor": "Fornecedor"},
                    "itens": [{"line_id": "oc-item-abc", "codigo": "SKU-1"}],
                },
                {
                    "id": "local-os",
                    "tipo": "os",
                    "numero": "200",
                    "data_criacao": "2026-07-19",
                    "status": "emitido",
                    "dados": {"cliente": "Cliente"},
                    "itens": [{"line_id": "os-item-def", "codigo": "SKU-2"}],
                },
            ], ensure_ascii=False), encoding="utf-8")

            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "ID Linha", "ACAO"])
            ws.append(["", "oc-item-abc", "concluir"])
            ws.append(["local-os", "", "excluir"])
            upload = io.BytesIO()
            wb.save(upload)
            upload.seek(0)

            with (
                patch.object(app_module, "HISTORICO_FILE", str(history_path)),
                patch.object(app_module, "login_enabled", return_value=False),
                patch.object(app_module.supabase_data, "enabled", return_value=False),
            ):
                response = app_module.app.test_client().post(
                    "/importar_baixa_documentos",
                    data={"arquivo_baixa_documentos": (upload, "baixas.xlsx")},
                    content_type="multipart/form-data",
                )

            rows = json.loads(history_path.read_text(encoding="utf-8"))
            self.assertEqual(302, response.status_code)
            self.assertEqual(["local-oc"], [row["id"] for row in rows])
            self.assertEqual("emitido", rows[0]["status"])
            self.assertEqual("concluido", rows[0]["itens"][0]["line_status"])

    def test_bulk_status_upload_returns_to_selected_management_tab(self):
        upload = io.BytesIO(b"fake spreadsheet")
        with (
            patch.object(app_module, "login_enabled", return_value=False),
            patch.object(app_module, "importar_baixas_documentos", return_value={
                "atualizados": 1,
                "excluidos": 0,
                "linhas_atualizadas": 0,
                "linhas_excluidas": 0,
                "ignorados": 0,
                "erros": [],
            }),
        ):
            response = app_module.app.test_client().post(
                "/importar_baixa_documentos",
                data={
                    "tipo_baixa_documentos": "os",
                    "next_tab": "gestao-os",
                    "arquivo_baixa_documentos": (upload, "baixas.xlsx"),
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(302, response.status_code)
        self.assertIn("tab=gestao-os", response.headers["Location"])

    def test_bulk_status_upload_can_delete_specific_os_line(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            history_path = Path(tmpdir) / "historico.json"
            history_path.write_text(json.dumps([{
                "id": "local-os",
                "tipo": "os",
                "numero": "200",
                "data_criacao": "2026-07-19",
                "status": "emitido",
                "dados": {"cliente": "Cliente"},
                "itens": [
                    {"line_id": "os-item-keep", "codigo": "SKU-1"},
                    {"line_id": "os-item-drop", "codigo": "SKU-2"},
                ],
            }], ensure_ascii=False), encoding="utf-8")

            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "ID Linha", "ACAO"])
            ws.append(["local-os", "os-item-drop", "excluir"])
            upload = io.BytesIO()
            wb.save(upload)
            upload.seek(0)

            with (
                patch.object(app_module, "HISTORICO_FILE", str(history_path)),
                patch.object(app_module, "login_enabled", return_value=False),
                patch.object(app_module.supabase_data, "enabled", return_value=False),
            ):
                response = app_module.app.test_client().post(
                    "/importar_baixa_documentos",
                    data={"arquivo_baixa_documentos": (upload, "baixas.xlsx")},
                    content_type="multipart/form-data",
                )

            rows = json.loads(history_path.read_text(encoding="utf-8"))
            self.assertEqual(302, response.status_code)
            self.assertEqual("emitido", rows[0]["status"])
            self.assertEqual(["os-item-keep"], [item["line_id"] for item in rows[0]["itens"]])

    def test_transient_import_path_is_scoped_by_login(self):
        with app_module.app.test_request_context("/"):
            app_module.session["suprimentos_user"] = {"id": 1, "username": "compras"}
            compras_path = app_module._user_scoped_file(app_module.OS_IMPORT_FILE)
            app_module.session["suprimentos_user"] = {"id": 2, "username": "producao"}
            producao_path = app_module._user_scoped_file(app_module.OS_IMPORT_FILE)

        self.assertNotEqual(compras_path, producao_path)
        self.assertIn("compras", compras_path)
        self.assertIn("producao", producao_path)

    def test_document_normalization_keeps_management_fields(self):
        row = supabase_data.normalizar_documento({
            "tipo": "OS",
            "numero": "42",
            "status": "CONCLUIDO",
            "submit_token": "abc",
            "criado_por": "maria",
            "atualizado_por": "joao",
            "dados": {},
        })
        self.assertEqual("os", row["tipo"])
        self.assertEqual("concluido", row["status"])
        self.assertEqual("abc", row["submit_token"])
        self.assertEqual("maria", row["criado_por"])
        self.assertEqual("joao", row["atualizado_por"])

    def test_document_normalization_keeps_immutable_erp_links(self):
        purchase_id = "11111111-1111-1111-1111-111111111111"
        work_id = "22222222-2222-2222-2222-222222222222"

        row = supabase_data.normalizar_documento({
            "tipo": "oc",
            "numero": "42",
            "erp_purchase_order_id": purchase_id,
            "erp_work_order_id": work_id,
            "dados": {},
        })
        legacy = supabase_data.documento_to_legacy(row)

        self.assertEqual(purchase_id, row["erp_purchase_order_id"])
        self.assertEqual(work_id, row["erp_work_order_id"])
        self.assertEqual(purchase_id, legacy["erp_purchase_order_id"])
        self.assertEqual(work_id, legacy["erp_work_order_id"])

    def test_erp_link_migration_is_additive_and_does_not_backfill_silently(self):
        migration_path = (
            Path(__file__).resolve().parents[1]
            / "supabase_suprimentos_erp_links_additive.sql"
        )
        sql = "\n".join(
            line for line in migration_path.read_text(encoding="utf-8").splitlines()
            if not line.lstrip().startswith("--")
        ).lower()

        self.assertIn("begin;", sql)
        self.assertIn("commit;", sql)
        self.assertIn("not valid", sql)
        self.assertNotIn("delete from", sql)
        self.assertNotIn("truncate", sql)
        self.assertNotIn("drop table", sql)
        self.assertNotIn("update public.suprimentos_documentos", sql)

    def test_reused_submission_token_does_not_overwrite_another_document(self):
        existing = [{
            "id": 10,
            "tipo": "os",
            "numero": "3099",
            "submit_token": "browser-token",
        }]
        saved = [{
            "id": 11,
            "tipo": "os",
            "numero": "3100",
            "submit_token": "browser-token::os::3100",
        }]
        with patch.object(supabase_data, "_request", side_effect=[existing, saved]) as request_mock:
            result = supabase_data.salvar_documento({
                "tipo": "os",
                "numero": "3100",
                "submit_token": "browser-token",
                "dados": {},
            })

        self.assertEqual(11, result["id"])
        self.assertEqual("browser-token::os::3100", result["submit_token"])
        post_call = request_mock.call_args_list[1]
        self.assertEqual("POST", post_call.args[0])
        self.assertEqual("browser-token::os::3100", post_call.kwargs["payload"]["submit_token"])
        self.assertIn(("on_conflict", "submit_token"), post_call.kwargs["query"])

    def test_same_submission_token_and_number_remain_idempotent(self):
        existing = [{
            "id": 10,
            "tipo": "os",
            "numero": "3100",
            "submit_token": "browser-token",
        }]
        saved = [{
            "id": 10,
            "tipo": "os",
            "numero": "3100",
            "submit_token": "browser-token",
        }]
        with patch.object(supabase_data, "_request", side_effect=[existing, saved]) as request_mock:
            result = supabase_data.salvar_documento({
                "tipo": "os",
                "numero": "3100",
                "submit_token": "browser-token",
                "dados": {},
            })

        self.assertEqual(10, result["id"])
        post_call = request_mock.call_args_list[1]
        self.assertEqual("browser-token", post_call.kwargs["payload"]["submit_token"])

    def test_document_history_has_no_default_quantity_limit(self):
        rows = [
            {
                "id": idx,
                "tipo": "os",
                "numero": str(idx),
                "data_criacao": "2026-07-27",
            }
            for idx in range(1005)
        ]
        with patch.object(supabase_data, "_all_rows", return_value=rows):
            documents = supabase_data.carregar_documentos()
            limited = supabase_data.carregar_documentos(limit=25)

        self.assertEqual(1005, len(documents))
        self.assertEqual(25, len(limited))

    def test_filtered_reports_include_status_users_and_details(self):
        history = [{
            "id": 12,
            "tipo": "oc",
            "numero": "100",
            "data_criacao": "2026-07-19",
            "status": "rascunho",
            "criado_por": "ana",
            "atualizado_por": "bia",
            "dados": {"fornecedor": "Fornecedor", "total_pedido": 123.45},
            "itens": [{"codigo": "SKU-1", "descricao": "Item", "qtd": 2, "total": 123.45}],
        }]
        with (
            patch.object(app_module, "login_enabled", return_value=False),
            patch.object(app_module, "carregar_historico", return_value=history),
        ):
            response = app_module.app.test_client().get("/exportar_dashboard?tipo=oc")

        self.assertEqual(200, response.status_code)
        workbook = load_workbook(io.BytesIO(response.data), read_only=True)
        self.assertEqual(["Compras", "Compras Itens"], workbook.sheetnames)
        main_rows = list(workbook["Compras"].iter_rows(values_only=True))
        item_rows = list(workbook["Compras Itens"].iter_rows(values_only=True))
        self.assertIn("Status", main_rows[0])
        self.assertIn("Criado Por", main_rows[0])
        self.assertIn("ACAO", main_rows[0])
        self.assertIn("ID Linha", item_rows[0])
        self.assertIn("Status Linha", item_rows[0])
        self.assertIn("ACAO", item_rows[0])
        self.assertEqual("rascunho", main_rows[1][1])
        self.assertEqual("SKU-1", item_rows[1][8])
        workbook.close()
        response.close()

    def test_os_report_includes_action_column_for_bulk_upload(self):
        history = [{
            "id": "os-1",
            "tipo": "os",
            "numero": "200",
            "data_criacao": "2026-07-19",
            "status": "emitido",
            "dados": {"cliente": "Cliente", "chassis": "ABC"},
            "itens": [{"line_id": "os-item-1", "codigo": "SKU-1", "descricao": "Item"}],
        }]
        with (
            patch.object(app_module, "login_enabled", return_value=False),
            patch.object(app_module, "carregar_historico", return_value=history),
        ):
            response = app_module.app.test_client().get("/exportar_dashboard?tipo=os")

        workbook = load_workbook(io.BytesIO(response.data), read_only=True)
        self.assertEqual(["Ordens de Servico", "OS Itens", "OS Processos", "OS Componentes"], workbook.sheetnames)
        self.assertIn("ACAO", [cell.value for cell in workbook["Ordens de Servico"][1]])
        self.assertIn("ID Linha", [cell.value for cell in workbook["OS Itens"][1]])
        self.assertIn("Status Linha", [cell.value for cell in workbook["OS Itens"][1]])
        self.assertIn("ACAO", [cell.value for cell in workbook["OS Itens"][1]])
        workbook.close()
        response.close()

    def test_purchase_can_be_saved_without_generating_document(self):
        with (
            patch.object(app_module, "login_enabled", return_value=False),
            patch.object(app_module, "atualizar_skus_automatico", return_value={}),
            patch.object(app_module, "carregar_fornecedores", return_value={}),
            patch.object(app_module, "carregar_produtos", return_value={}),
            patch.object(app_module, "carregar_os_componentes", return_value={}),
            patch.object(app_module, "proximo_numero_oc", return_value=51),
            patch.object(app_module, "registrar_historico") as register,
            patch.object(app_module, "gerar_word") as generate,
        ):
            response = app_module.app.test_client().post("/gerar_oc", data={
                "acao": "salvar",
                "oc_submit_token": "oc-token",
                "fornecedor": "Fornecedor",
                "codigo[]": "SKU-1",
                "descricao[]": "Item",
                "unidade[]": "UN",
                "qtd[]": "2",
                "valor[]": "10",
                "desconto[]": "0",
                "frete": "0",
            })

        self.assertEqual(302, response.status_code)
        generate.assert_not_called()
        self.assertEqual("rascunho", register.call_args.kwargs["status"])
        self.assertEqual("oc-token", register.call_args.kwargs["submit_token"])

    def test_emitted_purchase_edit_updates_same_erp_order_without_printing(self):
        existing = {
            "id": "doc-10",
            "tipo": "oc",
            "numero": "10",
            "status": "emitido",
            "dados": {"fornecedor": "Fornecedor"},
            "itens": [],
        }
        saved = {**existing, "dados": {"fornecedor": "Fornecedor atualizado"}}
        with (
            patch.object(app_module, "login_enabled", return_value=False),
            patch.object(app_module, "atualizar_skus_automatico", return_value={}),
            patch.object(app_module, "carregar_fornecedores", return_value={}),
            patch.object(app_module, "carregar_produtos", return_value={}),
            patch.object(app_module, "carregar_os_componentes", return_value={}),
            patch.object(app_module, "obter_historico_documento", return_value=existing),
            patch.object(
                app_module,
                "_sync_emitted_legacy_oc_to_erp",
                return_value={"id": "11111111-1111-1111-1111-111111111111", "updated": True},
            ) as sync,
            patch.object(app_module, "registrar_historico", return_value=saved) as register,
            patch.object(app_module, "vincular_documento_erp") as link,
            patch.object(app_module, "gerar_word") as generate,
        ):
            response = app_module.app.test_client().post("/gerar_oc", data={
                "acao": "salvar",
                "oc_historico_id": "doc-10",
                "oc_submit_token": "oc-token",
                "oc_numero": "10",
                "fornecedor": "Fornecedor atualizado",
                "codigo[]": "SKU-1",
                "descricao[]": "Item atualizado",
                "unidade[]": "UN",
                "qtd[]": "2",
                "valor[]": "10",
                "desconto[]": "0",
                "frete": "0",
                "oc_categoria": "BANCOS",
                "destino": "O.S. 712",
            })

        self.assertEqual(302, response.status_code)
        generate.assert_not_called()
        sync.assert_called_once()
        self.assertEqual("emitido", register.call_args.kwargs["status"])
        dados_sincronizados = sync.call_args.args[1]
        self.assertEqual("BANCOS", dados_sincronizados["oc_categoria"])
        self.assertEqual("O.S. 712", dados_sincronizados["destino"])
        dados_salvos = register.call_args.args[2]
        self.assertEqual("BANCOS", dados_salvos["oc_categoria"])
        self.assertEqual("O.S. 712", dados_salvos["destino"])
        link.assert_called_once_with(
            saved,
            "erp_purchase_order_id",
            "11111111-1111-1111-1111-111111111111",
        )

    def test_purchase_sync_uses_persisted_category_and_destination(self):
        history = {
            "id": "doc-10",
            "data_criacao": "2026-07-20",
        }
        purchase_data = {
            "previsao": "2026-08-01",
            "frete": 25,
            "obs": "Compra de bancos",
            "oc_categoria": "BANCOS",
            "destino": "O.S. 712",
        }
        items = [{
            "codigo": "SKU-1",
            "descricao": "Conjunto de bancos",
            "unidade": "UN",
            "qtd": 2,
            "valor": 100,
        }]

        with (
            patch.object(app_module, "erp_feature_enabled", return_value=True),
            patch.object(
                app_module,
                "_erp_stock_request",
                return_value={"id": "11111111-1111-1111-1111-111111111111"},
            ) as request_erp,
        ):
            app_module._sync_emitted_legacy_oc_to_erp(
                history,
                purchase_data,
                items,
                "10",
                "Fornecedor",
            )

        payload = request_erp.call_args.args[2]
        self.assertEqual("BANCOS", payload["categoria"])
        self.assertEqual("O.S. 712", payload["destino"])
        self.assertEqual("O.S. 712", payload["lines"][0]["destino"])
        self.assertEqual("suprimentos-oc:doc-10", payload["idempotency_key"])

    def test_purchase_history_api_attaches_same_integrated_order_for_editing(self):
        document = {
            "id": "doc-10",
            "tipo": "oc",
            "numero": "10",
            "dados": {"fornecedor": "Fornecedor"},
        }
        order = {
            "id": "11111111-1111-1111-1111-111111111111",
            "idempotency_key": "suprimentos-oc:doc-10",
            "categoria": "BANCOS",
            "destino": "O.S. 712",
        }
        with (
            patch.object(app_module, "login_enabled", return_value=False),
            patch.object(app_module, "obter_historico_documento", return_value=document),
            patch.object(app_module, "erp_feature_enabled", return_value=True),
            patch.object(app_module, "_erp_stock_request", return_value={"orders": [order]}),
        ):
            response = app_module.app.test_client().get("/api/historico/oc/doc-10")

        self.assertEqual(200, response.status_code)
        self.assertEqual(order, response.get_json()["documento"]["erp_purchase_order"])

    def test_purchase_editor_switches_tab_before_clearing_and_populating_form(self):
        template = (
            Path(__file__).resolve().parents[1]
            / "compras_app"
            / "templates"
            / "index.html"
        ).read_text(encoding="utf-8")
        start = template.index("function editarDocumentoHistoricoOC(doc)")
        end = template.index("async function editarHistoricoOC", start)
        editor = template[start:end]

        self.assertLess(editor.index("tabButton.click()"), editor.index("limparOCForm()"))
        self.assertEqual(1, editor.count("tabButton.click()"))

    def test_purchase_edit_is_not_overwritten_after_stock_receipt(self):
        existing = {
            "id": "doc-10",
            "tipo": "oc",
            "numero": "10",
            "status": "emitido",
            "dados": {"fornecedor": "Fornecedor"},
            "itens": [],
        }
        with (
            patch.object(app_module, "login_enabled", return_value=False),
            patch.object(app_module, "atualizar_skus_automatico", return_value={}),
            patch.object(app_module, "carregar_fornecedores", return_value={}),
            patch.object(app_module, "carregar_produtos", return_value={}),
            patch.object(app_module, "carregar_os_componentes", return_value={}),
            patch.object(app_module, "obter_historico_documento", return_value=existing),
            patch.object(
                app_module,
                "_sync_emitted_legacy_oc_to_erp",
                return_value={"id": "11111111-1111-1111-1111-111111111111", "locked": True},
            ),
            patch.object(app_module, "registrar_historico") as register,
            patch.object(app_module, "gerar_word") as generate,
        ):
            response = app_module.app.test_client().post("/gerar_oc", data={
                "acao": "salvar",
                "oc_historico_id": "doc-10",
                "oc_numero": "10",
                "fornecedor": "Fornecedor",
                "codigo[]": "SKU-1",
                "descricao[]": "Item",
                "unidade[]": "UN",
                "qtd[]": "2",
                "valor[]": "10",
                "desconto[]": "0",
                "frete": "0",
            })

        self.assertEqual(302, response.status_code)
        self.assertIn("gestao-oc", response.headers["Location"])
        register.assert_not_called()
        generate.assert_not_called()

    def test_submitter_action_is_preserved_before_buttons_are_disabled(self):
        template = (
            Path(__file__).resolve().parents[1]
            / "compras_app"
            / "templates"
            / "index.html"
        ).read_text(encoding="utf-8")

        self.assertIn("function preservarAcaoSubmitter", template)
        self.assertGreaterEqual(
            template.count("preservarAcaoSubmitter("),
            3,
        )

    def test_service_order_can_be_saved_without_generating_zip(self):
        with (
            patch.object(app_module, "login_enabled", return_value=False),
            patch.object(app_module, "atualizar_skus_automatico", return_value={}),
            patch.object(app_module, "carregar_os_produtos", return_value={}),
            patch.object(app_module, "carregar_produtos", return_value={}),
            patch.object(app_module, "carregar_regras_popup_item", return_value=[]),
            patch.object(app_module, "carregar_os_componentes", return_value={}),
            patch.object(app_module, "carregar_os_processos", return_value={}),
            patch.object(app_module, "carregar_relacoes_processo_item", return_value={}),
            patch.object(app_module, "proximo_numero_os", return_value=77),
            patch.object(app_module, "registrar_historico") as register,
            patch.object(app_module, "gerar_os_docx") as generate,
        ):
            response = app_module.app.test_client().post("/gerar_os", data={
                "acao": "salvar",
                "os_submit_token": "os-token",
                "os_composicao_json": "[]",
                "os_composicao_source": "custom",
                "os_cliente": "Cliente",
                "os_codigo[]": "",
            })

        self.assertEqual(302, response.status_code)
        generate.assert_not_called()
        self.assertEqual("rascunho", register.call_args.kwargs["status"])
        self.assertEqual("os-token", register.call_args.kwargs["submit_token"])

    def test_marco_zero_prefers_newest_duplicate_source(self):
        sources = [
            {
                "nome": "OS-antiga.docx",
                "conteudo": b"antiga",
                "data_arquivo": datetime(2026, 7, 20, 10, 0),
                "revisao": 0,
                "ordem": 0,
            },
            {
                "nome": "OS-R01.docx",
                "conteudo": b"nova",
                "data_arquivo": datetime(2026, 7, 21, 10, 0),
                "revisao": 1,
                "ordem": 1,
            },
        ]
        with patch.object(
            app_module,
            "parse_os_docx_atualizado",
            side_effect=[
                {"os_numero": "3090", "itens": [{"codigo": "A"}]},
                {"os_numero": "3090", "itens": [{"codigo": "B"}]},
            ],
        ):
            selected, discarded = app_module._parsear_fontes_os_reconciliacao(sources)

        self.assertEqual(1, discarded)
        self.assertEqual("OS-R01.docx", selected["3090"]["nome"])

    def test_marco_zero_recalculates_bom_and_keeps_required_choices(self):
        source = {
            "nome": "OS-3090.docx",
            "data_arquivo": datetime(2026, 7, 21, 10, 0),
            "dados": {
                "os_numero": "3090",
                "cliente": "Cliente",
                "chassis": "CHASSI",
                "itens": [
                    {"codigo": "40340001", "descricao": "CJ TETO", "qtd": 1, "unidade": "un"},
                    {"codigo": "FD-1", "descricao": "FATURAMENTO DIRETO SERVICO", "qtd": 1, "unidade": "un"},
                ],
                "composicao": [{"codigo": "10260092", "qtd": 2}],
            },
        }
        reference = {
            "tipo": "os",
            "numero": "3090",
            "data_criacao": "2026-07-20",
            "itens": [{"codigo": "FD-1", "fornecedor": "INSTALL TECH"}],
            "composicao": [],
        }
        products = {
            "40340001": {"descricao": "CJ TETO", "unidade": "un"},
            "FD-1": {"descricao": "FATURAMENTO DIRETO SERVICO", "unidade": "un"},
            "10260092": {"descricao": "LUMINARIA", "unidade": "un"},
            "COMP-1": {"descricao": "COMPONENTE", "unidade": "un"},
        }
        context = {
            "os_produtos": products,
            "produtos_catalogo": products,
            "componentes": {
                "40340001": [{"codigo": "COMP-1", "descricao": "COMPONENTE", "unidade": "un", "quantidade": 3}],
            },
            "processos": {},
            "processo_por_item": {},
            "regras_por_gatilho": {},
        }
        with (
            patch.object(app_module, "_resolver_nome_cliente_os", return_value="Cliente"),
            app_module.app.test_request_context("/"),
        ):
            document = app_module._montar_os_reconciliada(source, reference, context)

        composition_codes = {line["codigo"] for line in document["composicao"]}
        direct_item = next(item for item in document["itens"] if item["codigo"] == "FD-1")
        self.assertEqual({"COMP-1", "10260092", "FD-1"}, composition_codes)
        self.assertEqual("INSTALL TECH", direct_item["fornecedor"])
        self.assertTrue(all(line.get("line_id") for line in document["composicao"]))

    def test_selective_recalculation_preserves_new_service_orders(self):
        source = {
            "nome": "OS-3090.docx",
            "data_arquivo": datetime(2026, 7, 21, 10, 0),
            "dados": {"os_numero": "3090", "chassis": "BASE-1", "itens": [{"codigo": "SKU-1"}]},
        }
        baseline = {
            "id": 10,
            "tipo": "os",
            "numero": "3090",
            "data_criacao": "2026-07-21",
            "status": "emitido",
            "submit_token": "baseline-token",
            "criado_por": "paulo",
            "dados": {"chassis": "BASE-1"},
            "itens": [{"codigo": "SKU-1", "qtd": 10}],
            "processos": {},
            "composicao": [],
        }
        new_order = {
            "id": 11,
            "tipo": "os",
            "numero": "4000",
            "data_criacao": "2026-07-22",
            "status": "emitido",
            "dados": {"chassis": "NOVA-1"},
            "itens": [{"codigo": "SKU-NOVA", "qtd": 1}],
            "processos": {},
            "composicao": [],
        }
        corrected = {
            "tipo": "os",
            "numero": "3090",
            "data_criacao": "2026-07-21",
            "status": "emitido",
            "submit_token": "novo-token",
            "criado_por": "codex",
            "atualizado_por": "codex",
            "dados": {"chassis": "BASE-1"},
            "itens": [{"codigo": "SKU-1", "qtd": 1}],
            "processos": {},
            "composicao": [{"codigo": "COMP-1", "qtd": 2}],
        }
        final_baseline = {"id": 10, **corrected}
        final_rows = [final_baseline, new_order]

        with (
            patch.object(app_module.supabase_data, "enabled", return_value=True),
            patch.object(app_module, "_fontes_os_reconciliacao", return_value=[source]),
            patch.object(app_module, "_parsear_fontes_os_reconciliacao", return_value=({"3090": source}, 0)),
            patch.object(app_module, "_carregar_historico_local", return_value=[]),
            patch.object(app_module, "_contexto_reconciliacao_os", return_value={}),
            patch.object(app_module, "_montar_os_reconciliada", return_value=corrected),
            patch.object(app_module, "current_username", return_value="codex"),
            patch.object(
                app_module.supabase_data,
                "carregar_documentos",
                side_effect=[[baseline, new_order], final_rows],
            ),
            patch.object(app_module.supabase_data, "atualizar_documento", return_value=True) as update,
            patch.object(app_module.supabase_data, "excluir_documento") as delete_one,
            patch.object(app_module.supabase_data, "excluir_documentos") as delete_many,
        ):
            result = app_module.recalcular_os_importadas([object()])

        self.assertEqual(1, result["atualizadas"])
        self.assertEqual(1, result["preservadas"])
        self.assertEqual("10", update.call_args.args[0])
        self.assertEqual("3090", update.call_args.args[1]["numero"])
        delete_one.assert_not_called()
        delete_many.assert_not_called()

    def test_add_missing_service_orders_preserves_existing_numbers(self):
        source_existing = {
            "nome": "OS-3088.docx",
            "data_arquivo": datetime(2026, 7, 21, 10, 0),
            "dados": {"os_numero": "3088", "chassis": "BASE-1", "itens": [{"codigo": "SKU-1"}]},
        }
        source_missing = {
            "nome": "OS-3090.docx",
            "data_arquivo": datetime(2026, 7, 21, 11, 0),
            "dados": {"os_numero": "3090", "chassis": "BASE-2", "itens": [{"codigo": "SKU-2"}]},
        }
        existing = {
            "id": 10,
            "tipo": "os",
            "numero": "3088",
            "data_criacao": "2026-07-21",
            "dados": {"chassis": "BASE-1"},
            "itens": [{"codigo": "SKU-1"}],
            "processos": {},
            "composicao": [],
        }
        missing = {
            "tipo": "os",
            "numero": "3090",
            "data_criacao": "2026-07-21",
            "dados": {"chassis": "BASE-2"},
            "itens": [{"codigo": "SKU-2"}],
            "processos": {},
            "composicao": [],
        }
        saved = {"id": 11, **missing}

        with (
            patch.object(app_module.supabase_data, "enabled", return_value=True),
            patch.object(
                app_module,
                "_fontes_os_reconciliacao",
                return_value=[source_existing, source_missing],
            ),
            patch.object(
                app_module,
                "_parsear_fontes_os_reconciliacao",
                return_value=({"3088": source_existing, "3090": source_missing}, 0),
            ),
            patch.object(app_module, "_carregar_historico_local", return_value=[]),
            patch.object(app_module, "_contexto_reconciliacao_os", return_value={}),
            patch.object(app_module, "_montar_os_reconciliada", return_value=missing) as build,
            patch.object(
                app_module.supabase_data,
                "carregar_documentos",
                side_effect=[[existing], [existing, saved]],
            ),
            patch.object(
                app_module.supabase_data,
                "salvar_documentos",
                return_value=[saved],
            ) as insert,
            patch.object(app_module.supabase_data, "excluir_documentos") as delete_many,
        ):
            result = app_module.adicionar_os_ausentes([object()])

        self.assertEqual(1, result["inseridas"])
        self.assertEqual(1, result["ignoradas_existentes"])
        self.assertEqual(["3090"], result["numeros"])
        self.assertEqual(source_missing, build.call_args.args[0])
        self.assertEqual(["3090"], [row["numero"] for row in insert.call_args.args[0]])
        delete_many.assert_not_called()

    def test_refresh_open_service_orders_uses_current_bom_and_preserves_closed_orders(self):
        open_order = {
            "id": "open-1",
            "tipo": "os",
            "numero": "3100",
            "data_criacao": "2026-08-03",
            "status": "emitido",
            "submit_token": "open-token",
            "criado_por": "pcp",
            "dados": {"cliente": "Cliente A", "observacao": "manter"},
            "itens": [{"line_id": "os-item-root", "codigo": "ROOT-1", "descricao": "Conjunto", "qtd": 2, "unidade": "pc"}],
            "processos": {"CORTE": [{"line_id": "processo-1", "atividade": "Cortar"}]},
            "componentes": {"snapshot": "anterior"},
            "composicao": [{
                "line_id": "os-comp-comp-1",
                "item": "ROOT-1",
                "codigo": "COMP-1",
                "descricao": "Componente antigo",
                "qtd": 1,
                "unidade": "pc",
                "level": 0,
                "setor": "PREPARACAO",
                "setor_manual": True,
                "line_status": "pendente",
            }, {
                # Material included manually by PCP.  A refresh must update
                # its own B.O.M. rather than reset the O.S. to ROOT-1 only.
                "line_id": "os-comp-manual-root",
                "item": "MANUAL-ROOT",
                "codigo": "MANUAL-ROOT",
                "descricao": "Revestimento escolhido manualmente",
                "qtd": 1,
                "unidade": "cj",
                "level": 0,
                "line_status": "nao_aplicavel",
            }, {
                "line_id": "os-comp-manual-old",
                "item": "MANUAL-ROOT",
                "codigo": "MANUAL-OLD",
                "descricao": "Componente antigo manual",
                "qtd": 1,
                "unidade": "pc",
                "level": 1,
            }, {
                # Legacy popup choice: old documents had this contextual
                # material only in the composition, not in the item payload.
                "line_id": "os-comp-popup-root",
                "item": "ROOT-1",
                "codigo": "POPUP-ROOT",
                "descricao": "Opcao de popup mantida",
                "qtd": 1,
                "unidade": "pc",
                "level": 0,
            }, {
                "line_id": "os-comp-popup-old",
                "item": "POPUP-ROOT",
                "codigo": "POPUP-OLD",
                "descricao": "Componente antigo do popup",
                "qtd": 1,
                "unidade": "pc",
                "level": 1,
            }],
        }
        closed_order = {
            "id": "closed-1",
            "tipo": "os",
            "numero": "3099",
            "status": "concluido",
            "dados": {"cliente": "Cliente B"},
            "itens": [{"codigo": "ROOT-1", "qtd": 1}],
            "processos": {},
            "composicao": [{"codigo": "NAO-TOCAR"}],
        }
        purchase_order = {"id": "oc-1", "tipo": "oc", "numero": "2723", "status": "emitido"}
        store = {row["id"]: copy.deepcopy(row) for row in (open_order, closed_order, purchase_order)}

        def load_documents(*_args, **_kwargs):
            return copy.deepcopy(list(store.values()))

        def update_document(document_id, document):
            saved = copy.deepcopy(document)
            saved["id"] = str(document_id)
            store[str(document_id)] = saved
            return True

        with (
            patch.object(app_module.supabase_data, "enabled", return_value=True),
            patch.object(app_module.supabase_data, "carregar_documentos", side_effect=load_documents),
            patch.object(app_module.supabase_data, "atualizar_documento", side_effect=update_document) as update,
            patch.object(app_module, "carregar_os_componentes", return_value={
                "ROOT-1": [{"codigo": "COMP-1", "descricao": "Componente novo", "unidade": "pc", "quantidade": 3}],
                "MANUAL-ROOT": [{"codigo": "MANUAL-NEW", "descricao": "Componente novo manual", "unidade": "pc", "quantidade": 4}],
                "POPUP-ROOT": [{"codigo": "POPUP-NEW", "descricao": "Componente novo do popup", "unidade": "pc", "quantidade": 5}],
            }) as components,
            patch.object(app_module, "carregar_os_produtos", return_value={
                "ROOT-1": {"codigo": "ROOT-1", "descricao": "Conjunto", "unidade": "pc"},
                "COMP-1": {"codigo": "COMP-1", "descricao": "Componente novo", "unidade": "pc"},
                "MANUAL-ROOT": {"codigo": "MANUAL-ROOT", "descricao": "Revestimento escolhido manualmente", "unidade": "cj"},
                "MANUAL-NEW": {"codigo": "MANUAL-NEW", "descricao": "Componente novo manual", "unidade": "pc"},
                "POPUP-ROOT": {"codigo": "POPUP-ROOT", "descricao": "Opcao de popup mantida", "unidade": "pc"},
                "POPUP-NEW": {"codigo": "POPUP-NEW", "descricao": "Componente novo do popup", "unidade": "pc"},
            }) as catalog,
            patch.object(app_module, "current_username", return_value="pcp"),
        ):
            result = app_module.atualizar_bom_os_abertas()

        self.assertEqual(1, result["atualizadas"])
        self.assertEqual(1, result["encerradas"])
        self.assertEqual(1, update.call_count)
        components.assert_called_once_with(force=True)
        catalog.assert_called_once_with(force=True)
        refreshed = store["open-1"]
        self.assertEqual("COMP-1", refreshed["composicao"][0]["codigo"])
        self.assertEqual(6, refreshed["composicao"][0]["qtd"])
        self.assertEqual("os-comp-comp-1", refreshed["composicao"][0]["line_id"])
        self.assertEqual("pendente", refreshed["composicao"][0]["line_status"])
        self.assertEqual("PREPARACAO", refreshed["composicao"][0]["setor"])
        self.assertTrue(refreshed["composicao"][0]["setor_manual"])
        manual_root = next(linha for linha in refreshed["composicao"] if linha["codigo"] == "MANUAL-ROOT")
        manual_child = next(linha for linha in refreshed["composicao"] if linha["codigo"] == "MANUAL-NEW")
        self.assertEqual("nao_aplicavel", manual_root["line_status"])
        self.assertEqual(4, manual_child["qtd"])
        self.assertNotIn("MANUAL-OLD", [linha["codigo"] for linha in refreshed["composicao"]])
        self.assertIn("POPUP-ROOT", [linha["codigo"] for linha in refreshed["composicao"]])
        popup_child = next(linha for linha in refreshed["composicao"] if linha["codigo"] == "POPUP-NEW")
        self.assertEqual(5, popup_child["qtd"])
        self.assertNotIn("POPUP-OLD", [linha["codigo"] for linha in refreshed["composicao"]])
        self.assertEqual("anterior", refreshed["componentes"]["snapshot"])
        self.assertEqual("manter", refreshed["dados"]["observacao"])
        self.assertEqual("pcp", refreshed["dados"]["bom_atualizada_por"])
        self.assertEqual([{"codigo": "NAO-TOCAR"}], store["closed-1"]["composicao"])
        self.assertEqual("emitido", store["oc-1"]["status"])

    def test_refresh_open_service_orders_rolls_back_if_one_update_fails(self):
        first = {
            "id": "one",
            "tipo": "os",
            "numero": "3101",
            "status": "emitido",
            "dados": {},
            "itens": [{"codigo": "ROOT-1", "qtd": 1}],
            "processos": {},
            "composicao": [{"item": "ROOT-1", "codigo": "OLD-1", "qtd": 1}],
        }
        second = {
            "id": "two",
            "tipo": "os",
            "numero": "3102",
            "status": "emitido",
            "dados": {},
            "itens": [{"codigo": "ROOT-2", "qtd": 1}],
            "processos": {},
            "composicao": [{"item": "ROOT-2", "codigo": "OLD-2", "qtd": 1}],
        }
        store = {row["id"]: copy.deepcopy(row) for row in (first, second)}
        calls = []
        failed_once = {"value": False}

        def load_documents(*_args, **_kwargs):
            return copy.deepcopy(list(store.values()))

        def update_document(document_id, document):
            document_id = str(document_id)
            calls.append((document_id, copy.deepcopy(document)))
            if document_id == "two" and not failed_once["value"]:
                failed_once["value"] = True
                raise RuntimeError("indisponibilidade simulada")
            saved = copy.deepcopy(document)
            saved["id"] = document_id
            store[document_id] = saved
            return True

        with (
            patch.object(app_module.supabase_data, "enabled", return_value=True),
            patch.object(app_module.supabase_data, "carregar_documentos", side_effect=load_documents),
            patch.object(app_module.supabase_data, "atualizar_documento", side_effect=update_document),
            patch.object(app_module, "carregar_os_componentes", return_value={
                "ROOT-1": [{"codigo": "NEW-1", "quantidade": 1}],
                "ROOT-2": [{"codigo": "NEW-2", "quantidade": 1}],
            }),
            patch.object(app_module, "carregar_os_produtos", return_value={}),
            patch.object(app_module, "current_username", return_value="pcp"),
        ):
            with self.assertRaisesRegex(RuntimeError, "indisponibilidade simulada"):
                app_module.atualizar_bom_os_abertas()

        self.assertEqual([{"item": "ROOT-1", "codigo": "OLD-1", "qtd": 1}], store["one"]["composicao"])
        self.assertEqual([{"item": "ROOT-2", "codigo": "OLD-2", "qtd": 1}], store["two"]["composicao"])
        self.assertEqual(["one", "two", "one"], [document_id for document_id, _ in calls])


if __name__ == "__main__":
    unittest.main()
