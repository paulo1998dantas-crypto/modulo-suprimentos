import io
import os
import sys
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook


APP_DIR = Path(__file__).resolve().parents[1] / "compras_app"
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))
os.environ["SUPRIMENTOS_FILE_LOG"] = "0"

import app as app_module  # noqa: E402


class PcpNeedsReportTests(unittest.TestCase):
    def setUp(self):
        app_module.app.config.update(TESTING=True)
        self.client = app_module.app.test_client()

    def test_report_combines_open_os_needs_and_active_forecast_in_one_table(self):
        stock_projection = {
            "ok": True,
            "lines": [{
                "numero_os": "3100",
                "item_number": 3100,
                "chassi": "9VTESTE000000001",
                "cliente_nome": "Cliente O.S.",
                "codigo": "MP-001",
                "descricao": "Manta de teste",
                "unidade": "pc",
                "quantidade_necessaria": 5,
                "quantidade_coberta": 2,
                "quantidade_pendente": 3,
                "saldo_fluxo_compartilhado": 1,
                "setor": "PREPARAÇÃO",
                "itens_pai": "CJ-001",
                "status_necessidade": "PENDENTE",
                "data_entrega_os": "2026-08-15",
            }],
            "summary": {"work_orders": 1},
        }
        forecast_projection = [
            {
                "forecast_id": "forecast-1",
                "sku_codigo": "MP-001",
                "descricao": "Manta de teste",
                "unidade": "pc",
                "quantidade_planejada": 4,
                "origem": "BOM",
                "forecast": {
                    "codigo": "FCT-0001",
                    "tipo_demanda": "AGUARDANDO_CHEGADA",
                    "data_entrega_prevista": "2026-09-05",
                },
            },
            {
                "forecast_id": "forecast-2",
                "sku_codigo": "MP-001",
                "descricao": "Manta de teste",
                "unidade": "pc",
                "quantidade_planejada": 2,
                "origem": "BOM",
                "forecast": {
                    "codigo": "FCT-0002",
                    "tipo_demanda": "PREVISAO_DEMANDA",
                    "data_entrega_prevista": "2026-09-12",
                },
            },
        ]
        with (
            patch.object(app_module, "login_enabled", return_value=False),
            patch.object(app_module, "erp_feature_enabled", return_value=True),
            patch.object(app_module, "forecast_feature_enabled", return_value=True),
            patch.object(app_module, "can", return_value=True),
            patch.object(app_module, "_erp_stock_request", return_value=stock_projection) as stock_request,
            patch.object(
                app_module.supabase_data,
                "carregar_necessidades_forecasts_ativos",
                return_value=forecast_projection,
            ),
        ):
            response = self.client.get("/erp/relatorios/necessidades-pcp.xlsx")

        self.assertEqual(200, response.status_code)
        self.assertIn("Necessidades_PCP_", response.headers["Content-Disposition"])
        stock_request.assert_called_once_with("work-orders/needs")
        workbook = load_workbook(io.BytesIO(response.data), read_only=True)
        self.assertEqual(["Necessidades PCP"], workbook.sheetnames)
        values = list(workbook["Necessidades PCP"].iter_rows(values_only=True))
        header = values[3]
        self.assertIn("Origem", header)
        self.assertIn("Falta expedir / demanda planejada", header)
        rows = values[4:]
        os_row = next(row for row in rows if row[0] == "O.S.")
        confirmed_forecast = next(
            row for row in rows
            if row[0] == "FORECAST" and "CONFIRMADO" in row[2]
        )
        predictive_forecast = next(
            row for row in rows
            if row[0] == "FORECAST" and "PREDITIVO" in row[2]
        )
        self.assertEqual("3100", str(os_row[5]))
        self.assertEqual("2026-08-15", str(os_row[4]))
        self.assertEqual(3, os_row[14])
        self.assertEqual(4, confirmed_forecast[14])
        self.assertEqual(2, predictive_forecast[14])
        self.assertIsNone(confirmed_forecast[13])
        workbook.close()
        response.close()

    def test_management_screen_exposes_the_pcp_needs_export(self):
        template = (APP_DIR / "templates" / "erp_gestao_os.html").read_text(encoding="utf-8")
        self.assertIn("/erp/relatorios/necessidades-pcp.xlsx", template)
        self.assertIn("Entrega O.S.", template)
        with (
            patch.object(app_module, "login_enabled", return_value=False),
            patch.object(app_module, "erp_feature_enabled", return_value=True),
            patch.object(app_module, "can", return_value=True),
        ):
            response = self.client.get("/erp/gestao-os")
        self.assertEqual(200, response.status_code)
        self.assertIn(b"necessidades-pcp.xlsx", response.data)


if __name__ == "__main__":
    unittest.main()
